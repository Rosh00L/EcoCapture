from openpyxl import load_workbook
import numpy as np
import pandas as pd


# Load workbook
wb = load_workbook(r'C:\Github_proj\EcoCapture-Analytics\SQL scripts\photography_Keys.xlsx', read_only=True)

# Select a sheet
ws = wb['photokeys']

# Read data
li = []
for i in ws.iter_rows(values_only=True):
    li.append(i)
coef = np.array(li).ravel()
coef = np.array2string(coef,separator="|")

file1 = open('C:\Github_proj\EcoCapture-Analytics\QC files\photoKeys.txt', 'w')
file1.writelines(coef)
file1.close()

